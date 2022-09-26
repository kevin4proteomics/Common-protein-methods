import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy import stats
from sklearn import linear_model
from sklearn.metrics import mean_squared_error, r2_score
from pandas import DataFrame
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


class ProteinQuan():
    def standard_curve(dilution=1):
        """
        Input the concentrations of standard (e.g., BSA) in microgram per microliter and to corresponding A595 to
        get the standard curve
        :return: a plot of linear regression with the equation and R square printed on the plot.
        """
        conc = np.array([]) #empty arrays for data input
        asb = np.array([])

        sample_size = int(input('How many standards: '))
        for i in range(sample_size): #use for loop to inquire the concentration and A595 of standards sequentially
            i = i+1
            c = float(input('Concentration (ug/uL) for standard No.'+str(i)+': '))
            a = float(input('A595 for standard No.'+str(i)+': '))
            conc = np.append(conc,c)
            asb = np.append(asb,a)

        #calculate related statistics for linear regression
        slope, intercept, r_value, p_value, std_err = stats.linregress(conc,asb,)
        print("Slope = ", round(slope,3))
        print("Intercept = ", round(intercept,2))
        print("R-squared = ", round(r_value ** 2,2))
        print("Standard error = ", round(std_err,2))

        #draw a linear model
        regr = linear_model.LinearRegression()
        x = conc[:,np.newaxis]
        model=regr.fit(x,asb)
        yfit=model.predict(x)

        #generate annotation
        if intercept > 0:
            intercept = "+ " + str(round(intercept))
        else:
            intercept = "- " + str(round(abs(intercept)))
        print(intercept)
        r2 = r'$R^{2}$ = ' + str(round(r_value**2,3))
        equation = 'asb = ' + str(round(slope,2)) +' x conc (μg/μL) ' + str(intercept)
        fig, ax = plt.subplots(figsize=(5,5))
        ax.scatter(conc, asb, color='grey', alpha=0.5)
        ax.plot(x,yfit)

        #get coordinates for x and y position for text
        x0, xmax = plt.xlim()
        y0, ymax = plt.ylim()
        data_width = xmax - x0
        data_height = ymax-y0
        x_position = x0 + 0.05*data_width
        y_position = y0 + 0.93*data_height
        y_position_r = y0 + 0.88*data_height
        ax.text(x_position,y_position,s=equation, fontsize=12, fontstyle='normal')
        ax.text(x_position, y_position_r,r2, fontsize=12)
        ax.set_xlabel('Conc (μg/μL)', fontweight='bold')
        ax.set_ylabel('A595', fontweight='bold')
        plt.show()

    def bradford(path, slope, y_inter, loadingdye_conc=4):
        """
        This method converts UV-vis absorbance at 595 nm into protein concentration and generates an Excel table
        listing the sample information and recipe for preparing protein lysates for gel electrophoresis according to
        user's preference.
        :param slope: the slope of standard curve
        :param y_inter: y intercept of standard curve
        :param loadingdye_conc: the concentration of sample loading buffer, default is 4
        :return: an excel table showing the protein concentration and volume needed to prepare samples for SDS-PAGE
        """

        # setup directory
        # path = path + '/ProteinSample_mixture'  # directory to store information
        exp_name = str(input('Name of experiment: '))
        directory_name = exp_name + '_bradford'
        directory_path = str(path + '/'+directory_name)
        print('The data will be stored at : '+str(path))
        if os.path.exists(directory_path) == True:  # Create a new folder only when the directory is non-existing
            pass
        else:
            os.mkdir(directory_path)

        # Input basic data info
        sample_number = int(input('How many samples: '))
        volume_total = float(input('Sample volume (uL): '))

        # Create empty arrays for appending later
        samples = np.array([])
        asb = np.array([])
        dilution = np.array([])
        loading_dye = np.array([])
        total_volume = np.array([])
        f_conc = np.array([])

        # Input sample information
        while True:
            for i in range(sample_number):
                samples_input = str(input('Input name of sample No. '+str(i+1)+': ' ))
                asb_input = float(input('Input A595 asb of sample No. '+str(i+1)+': ' ))
                samples = np.append(samples, samples_input)
                asb = np.append(asb, asb_input)
            break

        while True:
            dilute_sample = str(input('Are samples further diluted from the way the standard curve was made (y/n)? '))
            if dilute_sample.lower() == 'y':
                while True:
                    dilution_question = input('Are all samples diluted in the same way (y/n)? ')
                    if dilution_question.lower() == 'y':
                        dilution_input = float(input('Samples were diluted by how many times compared to standard? '))
                        for i in range(sample_number):
                            dilution = np.append(dilution, dilution_input)
                        break

                    elif dilution_question.lower() == 'n':
                        while True:
                            for i in range(sample_number):
                                dilution_input = float(input('Dilution factor for sample No.'+str(i+1)+': '))
                                dilution = np.append(dilution, dilution_input)
                            if dilution.size == sample_number:
                                break
                        break
                    else:
                        print('Invalid input.')
                        # dilution_question = input('Are all samples diluted in the same way (y/n)? ')
                break
            elif dilute_sample.lower() == 'n':
                for i in range(sample_number):
                    dilution = np.append(dilution, 1)
                break
            else:
                print('Invalid input')

        # Calculate concentration
        conc = asb * dilution
        conc = np.round((conc-y_inter)/slope,2)
        lowest_amount = round(np.amin(conc) * volume_total, 2)
        print('Lowest amount: ' + str(lowest_amount) + ' ug')

        # Inquire how much protein should be used and the desired final conc
        total_protein = float(input('How much protein sample should be used (ug): '))
        max_conc = np.amin(conc)*(loadingdye_conc-1)/loadingdye_conc
        print('Max concentration: ' + str(round(max_conc, 2)) + ' (ug/uL)')
        final_conc = float(input('What is the desired final concentration (ug/ul): '))

        # Calculate the volume of protein samples, lysis, loading dye, and total
        total_volume_ind = round(total_protein / final_conc, 2)
        loading = total_volume_ind / 4

        protein_volume = np.round(total_protein / conc, 1)
        lysis = total_volume_ind * 0.75 - protein_volume
        lysis = np.round(lysis, 1)
        for i in range(sample_number):
            total_volume = np.append(total_volume, total_volume_ind)
            total_volume = np.round(total_volume, 1)
            loading_dye = np.append(loading_dye, loading)
            loading_dye = np.round(loading_dye, 1)
            f_conc = np.append(f_conc, final_conc)

        # output
        arr = np.array([samples, asb, dilution, conc, protein_volume, lysis, loading_dye, total_volume, f_conc])
        dframe1 = DataFrame(arr,
                            index=['Sample', 'A595', 'Dilution', 'Conc (ug/uL)', 'Protein Volume (uL)', 'Lysis (uL)',
                                   '4x Loading (uL)', 'Total Volume (uL)', 'Final Concentration (ug/uL)'])
        dframe1 = dframe1.T
        print(dframe1)
        file = str(directory_name+'/Bradford.xlsx')
        dframe1.to_excel(path+'/'+file)

        # open with openpyxl for formatting
        wb = load_workbook(filename=(path +'/'+ file))
        ws = wb.worksheets[0]
        ws.title = 'Bradford Assay'
        ws.delete_cols(1)

        ws.row_dimensions[1].height = 60

        # change column width
        dims = {}
        thin = Side(border_style="thin")

        for row in ws.rows:
            for cell in row:
                if cell.value:
                    cell.font = Font(size=14, name='Gill Sans MT')  # change fonts
                    dims[cell.column_letter] = 16
                    # dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 3))
                    # # +3 to make the column less
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    cell.border = Border(bottom=thin, top=thin)
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

        for cell in ws[1]:  # change the cell and font of the column header
            if cell.value:
                cell.font = Font(bold=True, size=14, name='Gill Sans MT')
                dims[cell.column_letter] = 19
        ws.column_dimensions['I'].width = 20
        wb.save(path+'/'+file)

